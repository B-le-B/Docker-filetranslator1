import os
import logging
import shutil
import tempfile
import re
import time
import hashlib
import threading
import copy
from typing import Optional, List, Dict, Any, Tuple, Union
from pathlib import Path
from tqdm import tqdm
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font, PatternFill, Border, Alignment, NamedStyle
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
# 修复Chart导入 - 使用具体的图表类型
from openpyxl.chart import (
    BarChart, LineChart, PieChart, ScatterChart, AreaChart, 
    DoughnutChart, RadarChart, BarChart3D, LineChart3D, PieChart3D
)

# 定义图表检查函数
def is_chart(obj):
    chart_types = (BarChart, LineChart, PieChart, ScatterChart, AreaChart, 
                  DoughnutChart, RadarChart, BarChart3D, LineChart3D, PieChart3D)
    return isinstance(obj, chart_types)

from openpyxl.comments import Comment
from openpyxl.formatting.rule import Rule
from concurrent.futures import ThreadPoolExecutor, as_completed, TimeoutError
from dataclasses import dataclass, field
from collections import OrderedDict
from contextlib import contextmanager
from enum import Enum
import datetime
import decimal

# 日志配置
logging.getLogger().setLevel(logging.INFO)

third_party_loggers = [
    'urllib3', 'requests', 'httpx', 'siliconflow', 'openai',
    'anthropic', 'zhipuai', 'dashscope', 'httpcore', 'httpx._client',
    'httpx._config', 'httpx._models', 'httpx._auth', 'requests.packages.urllib3',
    'requests_oauthlib', 'oauthlib', 'aiohttp', 'websockets', 'asyncio'
]

for logger_name in third_party_loggers:
    logging.getLogger(logger_name).setLevel(logging.WARNING)

class NoDebugFilter(logging.Filter):
    def filter(self, record):
        if record.levelno >= logging.WARNING:
            return True
        if record.name == __name__ or record.name.startswith('__main__'):
            return record.levelno >= logging.INFO
        return False

root_logger = logging.getLogger()
root_logger.addFilter(NoDebugFilter())

logger = logging.getLogger(__name__)

# ✅ 新增：多行文本处理常量
NEWLINE_PLACEHOLDER = "<<<EXCEL_NEWLINE_MARKER>>>"
CARRIAGE_RETURN_PLACEHOLDER = "<<<EXCEL_CR_MARKER>>>"
TAB_PLACEHOLDER = "<<<EXCEL_TAB_MARKER>>>"

# 预编译正则表达式
NUMBERED_LINE_PATTERN = re.compile(r'^\[(\d+)\]\s*(.*)')
FORMULA_PATTERN = re.compile(r'^=.+')
DATE_PATTERN = re.compile(r'^\d{4}[-/]\d{1,2}[-/]\d{1,2}|^\d{1,2}[-/]\d{1,2}[-/]\d{4}')
NUMBER_PATTERN = re.compile(r'^-?\d+\.?\d*%?$')
CODE_PATTERN = re.compile(r'^[A-Z]{2,}-?\d+|^\d+[A-Z]+\d*$', re.IGNORECASE)
CURRENCY_PATTERN = re.compile(r'[¥$€£₹]|\d+\.\d{2}$')

# ✅ 新增：文本转义处理函数
def _escape_excel_text_for_translation(text: str) -> str:
    """转义Excel文本中的特殊字符，防止破坏批次格式"""
    if not isinstance(text, str):
        return str(text)
    
    # 按顺序替换，避免冲突
    escaped = text
    escaped = escaped.replace('\r\n', CARRIAGE_RETURN_PLACEHOLDER + NEWLINE_PLACEHOLDER)  # Windows换行
    escaped = escaped.replace('\n\r', NEWLINE_PLACEHOLDER + CARRIAGE_RETURN_PLACEHOLDER)  # 少见格式
    escaped = escaped.replace('\r', CARRIAGE_RETURN_PLACEHOLDER)  # Mac换行
    escaped = escaped.replace('\n', NEWLINE_PLACEHOLDER)  # Unix换行
    escaped = escaped.replace('\t', TAB_PLACEHOLDER)  # 制表符
    
    return escaped

def _unescape_excel_text_from_translation(text: str) -> str:
    """还原翻译结果中的特殊字符"""
    if not isinstance(text, str):
        return str(text)
    
    # 按相反顺序还原
    unescaped = text
    unescaped = unescaped.replace(TAB_PLACEHOLDER, '\t')
    unescaped = unescaped.replace(NEWLINE_PLACEHOLDER, '\n')
    unescaped = unescaped.replace(CARRIAGE_RETURN_PLACEHOLDER, '\r')
    
    return unescaped

# 失败原因枚举
class FailureReason(Enum):
    """失败原因枚举"""
    TIMEOUT = "timeout"
    API_ERROR = "api_error"
    PARSE_ERROR = "parse_error"
    NOT_TRANSLATED = "not_translated"
    EMPTY_RESPONSE = "empty_response"
    CONNECTION_ERROR = "connection_error"
    BATCH_FAILURE = "batch_failure"
    COORDINATE_ERROR = "coordinate_error"
    SHEET_ERROR = "sheet_error"
    MAPPING_ERROR = "mapping_error"

@dataclass
class FailedTask:
    """失败任务数据类"""
    original_text: str
    original_index: int
    element: Any  # SafeExcelElement对象
    failure_reason: FailureReason
    retry_count: int = 0
    error_message: str = ""
    is_serious: bool = True
    
    def __post_init__(self):
        self.failure_timestamp = time.time()

@dataclass
class SafeExcelElement:
    """修复版安全Excel元素 - 精确映射定位系统"""
    # 核心内容
    full_text: str
    cell_value: Any
    
    # 精确定位信息
    sheet_name: str
    sheet_index: int
    row_index: int
    col_index: int
    cell_address: str  # "A1", "B2"等
    original_index: int  # ✅ 关键：收集时的原始顺序索引
    
    # 类型和格式信息
    cell_type: str = 'data'  # 'header', 'data', 'comment', 'chart_title'
    format_info: Dict[str, Any] = field(default_factory=dict)
    context_info: Dict[str, Any] = field(default_factory=dict)
    
    # 合并单元格信息
    is_merged: bool = False
    merged_range: Optional[str] = None  # "A1:B2"
    is_merge_top_left: bool = False
    
    # 特殊元素信息
    chart_info: Optional[Dict[str, Any]] = None
    comment_info: Optional[Dict[str, Any]] = None
    
    # 标识信息
    unique_id: str = ""
    element_type: str = 'excel_cell'
    
    def __post_init__(self):
        if not self.unique_id:
            # ✅ 唯一ID包含原始索引，确保精确映射
            self.unique_id = f"{self.sheet_name}!{self.cell_address}!{self.original_index}"

# Excel内容智能检测器
class ExcelContentDetector:
    """Excel内容智能检测器 - 增强版本"""
    
    @staticmethod
    def is_number(value: Any) -> bool:
        """检测是否为数字"""
        if isinstance(value, (int, float, decimal.Decimal)):
            return True
        if isinstance(value, str):
            return bool(NUMBER_PATTERN.match(value.strip()))
        return False
    
    @staticmethod
    def is_formula(value: Any) -> bool:
        """检测是否为Excel公式"""
        if isinstance(value, str):
            return bool(FORMULA_PATTERN.match(value.strip()))
        return False
    
    @staticmethod
    def is_date(value: Any) -> bool:
        """检测是否为日期"""
        if isinstance(value, datetime.datetime):
            return True
        if isinstance(value, str):
            return bool(DATE_PATTERN.match(value.strip()))
        return False
    
    @staticmethod
    def is_code_like(value: str) -> bool:
        """检测是否像代码/ID/SKU"""
        if not isinstance(value, str):
            return False
        value = value.strip()
        if len(value) < 3:
            return False
        return bool(CODE_PATTERN.match(value))
    
    @staticmethod
    def is_currency(value: Any) -> bool:
        """检测是否为货币"""
        if isinstance(value, str):
            return bool(CURRENCY_PATTERN.search(value))
        return False
    
    @staticmethod
    def is_url(value: str) -> bool:
        """检测是否为URL"""
        if not isinstance(value, str):
            return False
        value_upper = value.upper()
        return any(protocol in value_upper for protocol in ['HTTP://', 'HTTPS://', 'FTP://', 'FILE://'])
    
    @staticmethod
    def has_excessive_special_chars(value: str, threshold: float = 0.3) -> bool:
        """检测是否有过多特殊字符"""
        if not isinstance(value, str) or not value:
            return False
        special_count = sum(1 for c in value if c in '!@#$%^&*()_+-=[]{}|;:,.<>?')
        return (special_count / len(value)) > threshold
    
    @staticmethod
    def is_pure_whitespace_or_symbols(value: str) -> bool:
        """检测是否为纯空白字符或符号"""
        if not isinstance(value, str):
            return False
        # 检查是否只包含空白字符、标点符号、数字
        return not any(c.isalpha() for c in value)
    
    @staticmethod
    def should_translate(value: Any, cell_type: str = 'data', 
                        max_length: int = 200, min_alpha_ratio: float = 0.2) -> bool:
        """综合判断是否应该翻译 - 增强版本"""
        if not value:
            return False
        
        # 转为字符串处理
        text = str(value).strip()
        if not text:
            return False
        
        # 长度检查
        if len(text) > max_length or len(text) < 2:
            return False
        
        # 优先级检查 - 明确不翻译的类型
        if ExcelContentDetector.is_number(value):
            return False
        if ExcelContentDetector.is_formula(value):
            return False
        if ExcelContentDetector.is_date(value):
            return False
        if ExcelContentDetector.is_currency(value):
            return False
        if ExcelContentDetector.is_code_like(text):
            return False
        if ExcelContentDetector.is_url(text):
            return False
        if ExcelContentDetector.has_excessive_special_chars(text):
            return False
        if ExcelContentDetector.is_pure_whitespace_or_symbols(text):
            return False
        
        # 字母比例检查 - 确保有足够的文字内容
        alpha_count = sum(1 for c in text if c.isalpha())
        if len(text) > 3 and alpha_count / len(text) < min_alpha_ratio:
            return False
        
        # 数字比例检查 - 数字过多的不翻译
        digit_count = sum(1 for c in text if c.isdigit())
        if len(text) > 5 and digit_count / len(text) > 0.4:
            return False
        
        # 特殊字符检查 - 避免翻译配置文件或代码
        if text.count('=') > 1 or text.count(':') > 2:
            return False
        
        # 表头类型更宽松
        if cell_type in ['header', 'chart_title', 'axis_label', 'comment']:
            return True
        
        # 数据类型更严格检查
        if cell_type == 'data':
            # 避免翻译看起来像数据标识的内容
            if text.replace(' ', '').replace('-', '').replace('_', '').isalnum() and len(text) > 8:
                return False
            
            # 避免翻译版本号、序列号等
            if re.match(r'^v?\d+\.\d+', text.lower()):
                return False
        
        return True

# 智能LRU缓存
class SmartCache:
    """智能LRU缓存 - 增强版本"""
    
    def __init__(self, max_size: int = 1000):  # 进一步增加缓存大小
        self._cache = OrderedDict()
        self.max_size = max_size
        self._lock = threading.RLock()
        self._hits = 0
        self._misses = 0
        self._clears = 0
        self._overwrites = 0
    
    def get(self, key: str) -> Optional[str]:
        if not key or not isinstance(key, str):
            return None
            
        with self._lock:
            if key in self._cache:
                value = self._cache.pop(key)
                self._cache[key] = value  # 移到末尾
                self._hits += 1
                return value
            self._misses += 1
            return None
    
    def put(self, key: str, value: str):
        if not key or not isinstance(key, str) or not isinstance(value, str):
            return
            
        with self._lock:
            if key in self._cache:
                self._cache.pop(key)
                self._overwrites += 1
            elif len(self._cache) >= self.max_size:
                self._cache.popitem(last=False)  # 删除最旧的
            self._cache[key] = value
    
    def remove(self, key: str) -> bool:
        if not key or not isinstance(key, str):
            return False
            
        with self._lock:
            if key in self._cache:
                self._cache.pop(key)
                self._clears += 1
                return True
            return False
    
    def clear_pattern(self, pattern: str) -> int:
        if not pattern:
            return 0
            
        with self._lock:
            keys_to_remove = [key for key in self._cache.keys() if pattern in key]
            for key in keys_to_remove:
                self._cache.pop(key)
            self._clears += len(keys_to_remove)
            return len(keys_to_remove)
    
    def clear(self):
        with self._lock:
            cleared_count = len(self._cache)
            self._cache.clear()
            self._hits = 0
            self._misses = 0
            self._clears += cleared_count
            self._overwrites = 0
            
    @property
    def stats(self) -> Dict[str, Any]:
        with self._lock:
            total = self._hits + self._misses
            return {
                'size': len(self._cache),
                'max_size': self.max_size,
                'hits': self._hits,
                'misses': self._misses,
                'clears': self._clears,
                'overwrites': self._overwrites,
                'hit_rate': self._hits / total if total > 0 else 0,
                'utilization': len(self._cache) / self.max_size
            }

# 翻译验证器
class TranslationValidator:
    """翻译完整性验证器 - 增强版本"""
    
    ERROR_KEYWORDS = [
        'timeout', 'readtimeout', 'connecttimeout', 'httptimeout',
        'network error', 'connection error', 'api error', 'service error',
        'translation failed', 'service unavailable', 'request failed',
        'server error', 'bad gateway', 'gateway timeout', 'rate limit',
        '超时', '网络错误', '连接错误', '服务错误', '翻译失败',
        '服务不可用', '请求失败', '服务器错误', '限流', '频率限制'
    ]
    
    @staticmethod
    def is_error_message(text: str) -> bool:
        if not text:
            return True
            
        text_lower = text.lower()
        return any(keyword.lower() in text_lower for keyword in TranslationValidator.ERROR_KEYWORDS)
    
    @staticmethod
    def is_serious_failure(original_text: str, translated_text: str, 
                          large_text_threshold: int = 12, from_cache: bool = False) -> Tuple[bool, str]:
        """Excel版本的失败判断 - 精确阈值"""
        
        if from_cache:
            return False, "缓存结果"
        
        if not translated_text or translated_text.strip() == "":
            if len(original_text) > large_text_threshold:
                return True, f"大段文字未翻译（{len(original_text)}字符）"
            else:
                return False, f"短文本未翻译（{len(original_text)}字符，可能正常）"
        
        if TranslationValidator.is_error_message(translated_text):
            return True, f"翻译服务错误: {translated_text[:50]}..."
        
        if original_text.strip() == translated_text.strip():
            if len(original_text) > large_text_threshold:
                return True, f"长文本未翻译（{len(original_text)}字符）"
            else:
                return False, f"短文本与原文相同（{len(original_text)}字符，正常情况）"
        
        # 检查翻译是否被截断
        if len(translated_text) < len(original_text) * 0.3 and len(original_text) > 20:
            return True, f"翻译可能被截断（原文{len(original_text)}字符，译文{len(translated_text)}字符）"
        
        return False, "翻译成功"
    
    @staticmethod
    def validate_translation_quality(original_text: str, translated_text: str) -> Tuple[bool, str, float]:
        """验证翻译质量"""
        if not translated_text:
            return False, "翻译结果为空", 0.0
        
        # 计算质量分数
        score = 1.0
        issues = []
        
        # 长度合理性检查
        length_ratio = len(translated_text) / len(original_text) if original_text else 0
        if length_ratio < 0.3:
            score -= 0.4
            issues.append("翻译过短")
        elif length_ratio > 3.0:
            score -= 0.2
            issues.append("翻译过长")
        
        # 错误关键词检查
        if TranslationValidator.is_error_message(translated_text):
            score -= 0.6
            issues.append("包含错误信息")
        
        # 原文重复检查
        if original_text.strip() == translated_text.strip():
            score -= 0.3
            issues.append("与原文相同")
        
        is_good = score >= 0.6
        reason = "翻译质量良好" if is_good else f"翻译质量问题: {', '.join(issues)}"
        
        return is_good, reason, score

# 配置工具函数
def _prepare_prompt_config(prompt_config: Optional[Dict[str, Any]], kwargs: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """准备和标准化prompt配置 - 增强版本"""
    try:
        has_prompt_config = prompt_config and isinstance(prompt_config, dict) and prompt_config
        has_kwargs_config = any(k in kwargs for k in ['preserve_terms', 'glossary', 'additional_context', 'prompt_template', 'custom_prompt'])
        
        if not has_prompt_config and not has_kwargs_config:
            logger.debug("No valid prompt configuration found, returning None")
            return None
        
        config = {}
        if has_prompt_config:
            config.update(prompt_config)
        
        # 处理kwargs中的配置项
        for key in ['preserve_terms', 'glossary', 'additional_context', 'prompt_template', 'custom_prompt']:
            if key in kwargs and kwargs[key] is not None:
                config[key] = kwargs[key]
        
        if not config:
            logger.debug("Config is empty after processing, returning None")
            return None
        
        normalized_config = _normalize_prompt_config(config)
        
        if not normalized_config or normalized_config.get('mode') == 'none':
            logger.debug("Normalized config is invalid, returning None")
            return None
        
        logger.debug(f"Prepared prompt config for Excel translation: {normalized_config}")
        return normalized_config
        
    except Exception as e:
        logger.warning(f"Error preparing prompt config: {e}")
        return None

def _normalize_prompt_config(config: Dict[str, Any]) -> Dict[str, Any]:
    """标准化prompt配置格式 - 增强版本"""
    try:
        if not config or not isinstance(config, dict):
            return {'mode': 'none'}
        
        normalized = config.copy()
        
        # 确定模式
        if 'mode' not in normalized:
            if 'custom_prompt' in normalized and normalized['custom_prompt']:
                normalized['mode'] = 'custom'
            elif 'prompt_template' in normalized or 'professional_domain' in normalized:
                normalized['mode'] = 'professional'
            elif any(k in normalized for k in ['preserve_terms', 'glossary', 'additional_context']):
                normalized['mode'] = 'general'
            else:
                normalized['mode'] = 'none'
        
        # 处理保留术语
        preserve_terms = normalized.get('preserve_terms')
        if preserve_terms:
            try:
                if isinstance(preserve_terms, str):
                    # 支持多种分隔符
                    separators = [',', ';', '|', '\n']
                    terms_list = []
                    for sep in separators:
                        if sep in preserve_terms:
                            terms_list = [term.strip() for term in preserve_terms.split(sep) if term.strip()]
                            break
                    if not terms_list:  # 如果没有分隔符，作为单个术语
                        terms_list = [preserve_terms.strip()]
                    normalized['preserve_terms'] = terms_list
                elif isinstance(preserve_terms, list):
                    normalized['preserve_terms'] = [str(term).strip() for term in preserve_terms if str(term).strip()]
                else:
                    normalized.pop('preserve_terms', None)
            except Exception as e:
                logger.warning(f"Error processing preserve_terms: {e}")
                normalized.pop('preserve_terms', None)
        
        # 处理术语表
        glossary = normalized.get('glossary')
        if glossary:
            if not isinstance(glossary, dict):
                logger.warning(f"Glossary should be a dictionary, got {type(glossary)}, ignoring")
                normalized.pop('glossary', None)
            else:
                # 清理术语表，确保键值都是字符串
                clean_glossary = {}
                for k, v in glossary.items():
                    if k and v and isinstance(k, str) and isinstance(v, str):
                        clean_glossary[k.strip()] = v.strip()
                normalized['glossary'] = clean_glossary if clean_glossary else None
        
        # 处理自定义prompt
        if normalized.get('mode') == 'custom':
            custom_prompt = normalized.get('custom_prompt', {})
            if not custom_prompt or not isinstance(custom_prompt, dict):
                system_prompt = normalized.get('custom_system_prompt', normalized.get('system'))
                user_prompt = normalized.get('custom_user_prompt', normalized.get('user'))
                
                if system_prompt:
                    normalized['custom_prompt'] = {
                        'system': str(system_prompt),
                        'user': str(user_prompt) if user_prompt else 'Please translate the following Excel content to {target_lang}:\n\n{content}'
                    }
                else:
                    logger.warning("Custom mode selected but no valid custom prompt provided, falling back to general mode")
                    normalized['mode'] = 'general'
        
        # 处理专业模板
        if normalized.get('mode') == 'professional':
            domain = normalized.get('professional_domain', normalized.get('prompt_template', 'financial'))
            normalized['prompt_template'] = str(domain)
        
        # 清理空值
        normalized = {k: v for k, v in normalized.items() if v is not None}
        
        return normalized
        
    except Exception as e:
        logger.warning(f"Error normalizing prompt config: {e}")
        return {'mode': 'none'}

def _get_batch_settings_from_config(prompt_config: Optional[Dict[str, Any]], kwargs: Dict[str, Any]) -> Dict[str, int]:
    """从配置中获取批处理设置 - 精确版本"""
    settings = {
        'batch_size': 50,            # 用户要求的50
        'max_chars': 4000,           # 适应更大批次
        'max_workers': 5,            # 用户要求的5个线程
        'retry_max_workers': 5,
        'translation_timeout': 35,   # 适当增加超时时间
        'max_retries': 5,
        'large_text_threshold': 12,  # 用户调整的阈值
        'retry_failure_threshold': 0.0,
        'non_ascii_threshold': 0.0
    }
    
    try:
        # 从kwargs获取设置
        if 'batch_size' in kwargs and isinstance(kwargs['batch_size'], (int, float)):
            settings['batch_size'] = max(5, min(int(kwargs['batch_size']), 100))
        if 'max_chunk_size' in kwargs and isinstance(kwargs['max_chunk_size'], (int, float)):
            settings['max_chars'] = max(500, min(int(kwargs['max_chunk_size']), 15000))
        if 'max_workers' in kwargs and isinstance(kwargs['max_workers'], (int, float)):
            settings['max_workers'] = max(1, min(int(kwargs['max_workers']), 8))
        if 'retry_max_workers' in kwargs and isinstance(kwargs['retry_max_workers'], (int, float)):
            settings['retry_max_workers'] = max(1, min(int(kwargs['retry_max_workers']), 4))
        if 'translation_timeout' in kwargs and isinstance(kwargs['translation_timeout'], (int, float)):
            settings['translation_timeout'] = max(20, min(int(kwargs['translation_timeout']), 120))
        if 'max_retries' in kwargs and isinstance(kwargs['max_retries'], (int, float)):
            settings['max_retries'] = max(2, min(int(kwargs['max_retries']), 8))
        if 'large_text_threshold' in kwargs and isinstance(kwargs['large_text_threshold'], (int, float)):
            settings['large_text_threshold'] = max(5, min(int(kwargs['large_text_threshold']), 100))
        
        # 从prompt_config获取设置
        if prompt_config and isinstance(prompt_config, dict):
            max_units = prompt_config.get('max_units_per_chunk')
            if max_units and isinstance(max_units, (int, float)):
                settings['batch_size'] = max(5, min(int(max_units), 80))
            
            max_chars = prompt_config.get('max_chars_per_chunk')
            if max_chars and isinstance(max_chars, (int, float)):
                settings['max_chars'] = max(500, min(int(max_chars), 15000))
        
        # 验证设置合理性
        if settings['batch_size'] * 100 > settings['max_chars']:
            settings['max_chars'] = settings['batch_size'] * 120  # 确保字符数足够
            
        logger.debug(f"Batch settings: {settings}")
    
    except Exception as e:
        logger.warning(f"Error processing batch settings: {e}")
    
    return settings

def _validate_excel_file(filepath: str) -> Tuple[bool, str]:
    """验证Excel文件"""
    try:
        if not os.path.exists(filepath):
            return False, f"文件不存在: {filepath}"
        
        if not filepath.lower().endswith(('.xlsx', '.xls')):
            return False, f"不是Excel文件: {filepath}"
        
        file_size = os.path.getsize(filepath)
        if file_size == 0:
            return False, f"文件为空: {filepath}"
        
        if file_size > 100 * 1024 * 1024:  # 100MB
            return False, f"文件过大: {file_size / (1024*1024):.1f}MB，超过100MB限制"
        
        # 尝试加载文件
        try:
            workbook = load_workbook(filepath, read_only=True)
            sheet_count = len(workbook.sheetnames)
            workbook.close()
            
            if sheet_count == 0:
                return False, "Excel文件没有工作表"
            
            return True, f"Excel文件验证通过，包含{sheet_count}个工作表"
            
        except Exception as e:
            return False, f"Excel文件损坏或格式不支持: {e}"
        
    except Exception as e:
        return False, f"文件验证异常: {e}"

class AdvancedExcelTranslator:
    """高级Excel翻译器 - 精确映射定位版本"""
    
    def __init__(
        self, 
        translator, 
        batch_size: int = 50,                    # 用户要求的50
        max_chars: int = 4000,                   # 适应更大批次
        max_workers: int = 5,                    # 用户要求的5个线程
        retry_max_workers: int = 5,
        prompt_config: Optional[Dict[str, Any]] = None,
        reference_doc: Optional[str] = None,
        translation_timeout: int = 35,           # 适当增加超时
        max_retries: int = 5,
        large_text_threshold: int = 12,          # 用户调整到12
        retry_failure_threshold: float = 0.0,
        non_ascii_threshold: float = 0.0,
        # Excel特有参数
        translate_headers: bool = True,
        translate_comments: bool = True,
        translate_charts: bool = True,
        skip_formulas: bool = True,
        skip_numbers: bool = True,
        skip_dates: bool = True,
        max_text_length: int = 200,
        header_detection_rows: int = 3,          # 减少到3行
        min_alpha_ratio: float = 0.2,            # 用户要求的0.2
        selected_sheets: Optional[List[str]] = None,
        preserve_formulas: bool = True,
        preserve_data_validation: bool = True,
        smart_detection: bool = True,
        **kwargs
    ):
        self.translator = translator
        self.batch_size = batch_size
        self.max_chars = max_chars
        self.max_workers = max_workers
        self.retry_max_workers = retry_max_workers
        self.translation_timeout = translation_timeout
        self.max_retries = max_retries
        self.large_text_threshold = large_text_threshold
        self.retry_failure_threshold = retry_failure_threshold
        self.non_ascii_threshold = non_ascii_threshold
        
        # Excel特有设置
        self.translate_headers = translate_headers
        self.translate_comments = translate_comments
        self.translate_charts = translate_charts
        self.skip_formulas = skip_formulas
        self.skip_numbers = skip_numbers
        self.skip_dates = skip_dates
        self.max_text_length = max_text_length
        self.header_detection_rows = header_detection_rows
        self.min_alpha_ratio = min_alpha_ratio
        self.selected_sheets = selected_sheets or []
        self.preserve_formulas = preserve_formulas
        self.preserve_data_validation = preserve_data_validation
        self.smart_detection = smart_detection
        
        # 验证worker配置
        if self.retry_max_workers < 1:
            self.retry_max_workers = 1
        if self.retry_max_workers > 4:
            logger.warning(f"retry_max_workers={self.retry_max_workers}可能过高，建议不超过4")
        
        # Worker资源管理 - 适应5个线程
        total_max_workers = 10
        if self.max_workers + self.retry_max_workers > total_max_workers:
            self.max_workers = max(1, total_max_workers - self.retry_max_workers)
            logger.info(f"调整主翻译workers为{self.max_workers}，为重试预留{self.retry_max_workers}个workers")
        
        self.cache = SmartCache(1000)  # 增加缓存大小
        self.source_lang = None
        self.reference_doc = reference_doc
        self._config_lock = threading.RLock()
        self._cached_config_hash = None
        self.workbook = None  # ✅ 保持workbook引用用于精确定位
        
        # 失败任务追踪
        self.failed_tasks: List[FailedTask] = []
        self.failed_tasks_lock = threading.Lock()
        
        # 重试策略 - 适应更大批次
        self.retry_batch_sizes = [10, 8, 5, 3, 1]  # 更合理的重试批次
        self.retry_delays = [1, 2, 3, 4, 5]
        
        # 处理prompt配置
        try:
            self.effective_prompt_config = _prepare_prompt_config(prompt_config, kwargs)
            self.original_translator_config = None
            
            if self.effective_prompt_config is None:
                self.effective_prompt_config = {'mode': 'none'}
            
        except Exception as e:
            logger.warning(f"Error initializing prompt config: {e}")
            self.effective_prompt_config = {'mode': 'none'}
        
        # 从配置获取批处理设置
        try:
            batch_settings = _get_batch_settings_from_config(self.effective_prompt_config, kwargs)
            self.batch_size = batch_settings['batch_size']
            self.max_chars = batch_settings['max_chars']
            self.max_workers = min(self.max_workers, batch_settings['max_workers'])
            self.retry_max_workers = min(self.retry_max_workers, batch_settings['retry_max_workers'])
            self.translation_timeout = batch_settings['translation_timeout']
            self.max_retries = batch_settings['max_retries']
            self.large_text_threshold = batch_settings['large_text_threshold']
            self.retry_failure_threshold = batch_settings['retry_failure_threshold']
            self.non_ascii_threshold = batch_settings['non_ascii_threshold']
        except Exception as e:
            logger.warning(f"Error getting batch settings: {e}")
        
        if self.effective_prompt_config and self.effective_prompt_config.get('mode') != 'none':
            logger.info(f"AdvancedExcelTranslator initialized with prompt config: mode={self.effective_prompt_config.get('mode')}")
        
        # ✅ 增强的统计信息
        self.stats = {
            'total_sheets': 0,
            'processed_sheets': 0,
            'total_cells': 0,
            'translated_cells': 0,
            'skipped_cells': 0,
            'total_chars': 0,
            'total_batches': 0,
            'headers': 0,
            'data_cells': 0,
            'comments': 0,
            'chart_elements': 0,
            'formulas_preserved': 0,
            'numbers_skipped': 0,
            'dates_skipped': 0,
            'processing_time': 0,
            'prompt_mode': self.effective_prompt_config.get('mode', 'none'),
            'api_calls': 0,
            'cache_savings': 0,
            'serious_failures': 0,
            'minor_issues': 0,
            'retry_attempts': 0,
            'final_failures': 0,
            'cache_clears': 0,
            'final_rescues': 0,
            'retry_workers_used': 0,
            'concurrent_retry_batches': 0,
            'smart_detections': 0,
            'content_filters': 0,
            'coordinate_errors': 0,
            'merge_cell_handled': 0,
            'mapping_errors': 0,  # ✅ 新增映射错误统计
            'validation_failures': 0,  # ✅ 新增验证失败统计
            'precise_mappings': 0,  # ✅ 新增精确映射统计
        }

    @contextmanager
    def _translator_config_context(self):
        """线程安全的配置上下文管理器"""
        try:
            self._apply_prompt_config_to_translator()
            yield
        except Exception as e:
            logger.error(f"Error in translator config context: {e}")
            raise
        finally:
            self._restore_translator_config()

    def _apply_prompt_config_to_translator(self):
        """应用prompt配置到翻译器"""
        if (self.effective_prompt_config and 
            self.effective_prompt_config.get('mode') != 'none' and 
            hasattr(self.translator, 'set_prompt_config')):
            try:
                with self._config_lock:
                    self.original_translator_config = getattr(self.translator, 'prompt_config', None)
                    config_copy = copy.deepcopy(self.effective_prompt_config)
                    self.translator.set_prompt_config(config_copy)
                    logger.info("Applied prompt config to translator")
                    return True
            except Exception as e:
                logger.warning(f"Failed to apply prompt config to translator: {e}")
        return False

    def _restore_translator_config(self):
        """恢复翻译器的原始配置"""
        if (self.effective_prompt_config and 
            self.effective_prompt_config.get('mode') != 'none' and 
            hasattr(self.translator, 'set_prompt_config')):
            try:
                with self._config_lock:
                    if self.original_translator_config is not None:
                        self.translator.set_prompt_config(self.original_translator_config)
                    else:
                        if hasattr(self.translator, 'prompt_config'):
                            self.translator.prompt_config = None
                    logger.debug("Restored translator config")
            except Exception as e:
                logger.warning(f"Failed to restore translator config: {e}")

    def _get_config_safe(self) -> Optional[Dict[str, Any]]:
        """线程安全地获取配置副本"""
        with self._config_lock:
            return copy.deepcopy(self.effective_prompt_config) if self.effective_prompt_config else None

    def _collect_excel_elements_precisely(self) -> List[SafeExcelElement]:
        """精确收集Excel元素 - 保持原始顺序的映射系统"""
        elements = []
        original_index = 0  # ✅ 全局原始索引计数器
        
        try:
            sheet_names = self.workbook.sheetnames
            self.stats['total_sheets'] = len(sheet_names)
            
            logger.info(f"开始精确收集Excel元素，共{len(sheet_names)}个工作表")
            
            # ✅ 按工作表顺序收集，保持原始顺序
            for sheet_index, sheet_name in enumerate(sheet_names):
                # 检查是否选择翻译此工作表
                if self.selected_sheets and sheet_name not in self.selected_sheets:
                    logger.info(f"跳过工作表: {sheet_name} (未在选择列表中)")
                    continue
                
                logger.info(f"收集工作表元素: {sheet_name}")
                worksheet = self.workbook[sheet_name]
                self.stats['processed_sheets'] += 1
                
                # ✅ 收集工作表元素 - 传递并更新original_index
                sheet_elements, original_index = self._extract_elements_from_worksheet_precisely(
                    worksheet, sheet_name, sheet_index, original_index
                )
                elements.extend(sheet_elements)
                
                # 处理批注（如果启用）
                if self.translate_comments:
                    comment_elements, original_index = self._extract_comments_precisely(
                        worksheet, sheet_name, sheet_index, original_index
                    )
                    elements.extend(comment_elements)
                
                # 处理图表（如果启用）
                if self.translate_charts:
                    chart_elements, original_index = self._extract_charts_precisely(
                        worksheet, sheet_name, sheet_index, original_index
                    )
                    elements.extend(chart_elements)
        
        except Exception as e:
            logger.error(f"Error collecting Excel elements precisely: {e}")
        
        logger.info(f"精确收集完成：{len(elements)}个元素，最大索引：{original_index-1}")
        return elements

    def _extract_elements_from_worksheet_precisely(self, worksheet: Worksheet, sheet_name: str, 
                                                 sheet_index: int, start_index: int) -> Tuple[List[SafeExcelElement], int]:
        """从工作表中精确提取元素 - 保持原始顺序"""
        elements = []
        current_index = start_index
        
        try:
            # 获取工作表的实际使用范围
            if worksheet.max_row == 1 and worksheet.max_column == 1:
                # 检查是否真的是空工作表
                cell_a1 = worksheet.cell(row=1, column=1)
                if cell_a1.value is None:
                    logger.debug(f"工作表 {sheet_name} 为空")
                    return elements, current_index
            
            max_row = min(worksheet.max_row, 10000)
            max_col = min(worksheet.max_column, 100)
            
            logger.debug(f"工作表 {sheet_name} 范围: {max_row} 行 x {max_col} 列")
            
            # ✅ 收集合并单元格信息
            merged_ranges = self._get_merged_ranges_info(worksheet)
            
            # ✅ 按自然行列顺序遍历，保持原始索引递增
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    try:
                        cell = worksheet.cell(row=row, column=col)
                        cell_address = cell.coordinate
                        
                        if cell.value is not None:
                            self.stats['total_cells'] += 1
                            
                            # 检查是否为合并单元格的非左上角位置
                            if cell_address in merged_ranges and not merged_ranges[cell_address].get('is_top_left', True):
                                logger.debug(f"跳过合并单元格非左上角: {cell_address}")
                                continue
                            
                            # 智能检测是否应该翻译
                            if self._should_translate_cell_precisely(cell, row, col, worksheet):
                                # ✅ 创建元素时传入当前索引
                                element = self._create_precise_excel_element(
                                    cell, sheet_name, sheet_index, worksheet, merged_ranges, current_index
                                )
                                if element:
                                    elements.append(element)
                                    current_index += 1  # ✅ 递增索引
                                    self.stats['smart_detections'] += 1
                                    self.stats['precise_mappings'] += 1
                            else:
                                self.stats['skipped_cells'] += 1
                                self.stats['content_filters'] += 1
                                
                    except Exception as e:
                        logger.debug(f"处理单元格 {row},{col} 异常: {e}")
                        continue
        
        except Exception as e:
            logger.error(f"处理工作表 {sheet_name} 异常: {e}")
        
        logger.info(f"工作表 {sheet_name}: 收集到 {len(elements)} 个元素，索引范围 {start_index}-{current_index-1}")
        return elements, current_index


    def _get_merged_ranges_info(self, worksheet: Worksheet) -> Dict[str, Dict[str, Any]]:
        """获取合并单元格信息"""
        merged_ranges = {}
        
        try:
            for merged_range in worksheet.merged_cells.ranges:
                # 左上角单元格
                top_left_addr = f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}"
                merged_ranges[top_left_addr] = {
                    'range': str(merged_range),
                    'is_top_left': True,
                    'min_row': merged_range.min_row,
                    'min_col': merged_range.min_col,
                    'max_row': merged_range.max_row,
                    'max_col': merged_range.max_col
                }
                
                # 标记合并范围内的其他单元格
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        cell_addr = f"{get_column_letter(col)}{row}"
                        if cell_addr != top_left_addr:
                            merged_ranges[cell_addr] = {
                                'range': str(merged_range),
                                'is_top_left': False,
                                'top_left': top_left_addr
                            }
                            
        except Exception as e:
            logger.debug(f"获取合并单元格信息异常: {e}")
        
        return merged_ranges

    def _should_translate_cell_precisely(self, cell: Cell, row: int, col: int, worksheet: Worksheet) -> bool:
        """精确判断单元格是否应该翻译"""
        try:
            if cell.value is None:
                return False
            
            # 确定单元格类型
            cell_type = self._determine_cell_type_precisely(cell, row, col, worksheet)
            
            # 使用增强的智能检测器
            if self.smart_detection:
                should_translate = ExcelContentDetector.should_translate(
                    cell.value, 
                    cell_type=cell_type,
                    max_length=self.max_text_length,
                    min_alpha_ratio=self.min_alpha_ratio
                )
                return should_translate
            
            # 基础检测逻辑
            if isinstance(cell.value, str):
                text = cell.value.strip()
                if len(text) > self.max_text_length or len(text) < 2:
                    return False
                return True
            
            return False
            
        except Exception as e:
            logger.debug(f"检测单元格翻译性异常: {e}")
            return False

    def _determine_cell_type_precisely(self, cell: Cell, row: int, col: int, worksheet: Worksheet) -> str:
        """精确确定单元格类型"""
        try:
            # 检查是否为公式
            if self.skip_formulas and hasattr(cell, 'data_type') and cell.data_type == 'f':
                self.stats['formulas_preserved'] += 1
                return 'formula'
            
            # 检查是否为表头（在前几行）
            if self.translate_headers and row <= self.header_detection_rows:
                if self._looks_like_header_precisely(cell.value, row, col, worksheet):
                    self.stats['headers'] += 1
                    return 'header'
            
            # 检查是否为数字
            if self.skip_numbers and ExcelContentDetector.is_number(cell.value):
                self.stats['numbers_skipped'] += 1
                return 'number'
            
            # 检查是否为日期
            if self.skip_dates and ExcelContentDetector.is_date(cell.value):
                self.stats['dates_skipped'] += 1
                return 'date'
            
            # 默认为数据单元格
            self.stats['data_cells'] += 1
            return 'data'
            
        except Exception as e:
            logger.debug(f"确定单元格类型异常: {e}")
            return 'data'

    def _looks_like_header_precisely(self, value: Any, row: int, col: int, worksheet: Worksheet) -> bool:
        """精确判断是否像表头"""
        try:
            if not isinstance(value, str):
                return False
            
            text = value.strip()
            if len(text) < 2 or len(text) > 50:
                return False
            
            # 表头通常不包含大量数字
            digit_ratio = sum(1 for c in text if c.isdigit()) / len(text)
            if digit_ratio > 0.5:
                return False
            
            # 表头通常包含字母
            alpha_ratio = sum(1 for c in text if c.isalpha()) / len(text)
            if alpha_ratio < 0.3:
                return False
            
            # ✅ 增强的表头检测：检查是否在表格的第一行或第一列
            is_first_row = row == 1
            is_first_few_rows = row <= self.header_detection_rows
            
            # 检查同行或同列是否有其他类似的文本（表头通常成组出现）
            has_sibling_headers = False
            try:
                # 检查同行的其他单元格
                for check_col in range(max(1, col-2), min(worksheet.max_column+1, col+3)):
                    if check_col != col:
                        check_cell = worksheet.cell(row=row, column=check_col)
                        if (check_cell.value and isinstance(check_cell.value, str) and 
                            len(str(check_cell.value).strip()) > 1):
                            has_sibling_headers = True
                            break
            except:
                pass
            
            return (is_first_few_rows and alpha_ratio >= 0.3) or (has_sibling_headers and alpha_ratio >= 0.4)
            
        except Exception as e:
            logger.debug(f"表头判断异常: {e}")
            return False

    def _create_precise_excel_element(self, cell: Cell, sheet_name: str, sheet_index: int, 
                                    worksheet: Worksheet, merged_ranges: Dict[str, Any], 
                                    original_index: int) -> Optional[SafeExcelElement]:
        """创建精确映射的Excel元素"""
        try:
            cell_address = cell.coordinate
            row_index = cell.row
            col_index = cell.column
            
            # 检查合并单元格信息
            is_merged = cell_address in merged_ranges
            merged_range = None
            is_merge_top_left = False
            
            if is_merged:
                merge_info = merged_ranges[cell_address]
                merged_range = merge_info['range']
                is_merge_top_left = merge_info.get('is_top_left', False)
                if is_merge_top_left:
                    self.stats['merge_cell_handled'] += 1
            
            # 提取格式信息
            format_info = self._extract_cell_format_precisely(cell)
            
            # 构建上下文信息
            context_info = self._get_surrounding_context_precisely(worksheet, row_index, col_index)
            
            # ✅ 创建SafeExcelElement，传入original_index
            element = SafeExcelElement(
                full_text=str(cell.value),
                cell_value=cell.value,
                sheet_name=sheet_name,
                sheet_index=sheet_index,
                row_index=row_index,
                col_index=col_index,
                cell_address=cell_address,
                original_index=original_index,  # ✅ 关键：传入原始索引
                cell_type=self._determine_cell_type_precisely(cell, row_index, col_index, worksheet),
                format_info=format_info,
                context_info=context_info,
                is_merged=is_merged,
                merged_range=merged_range,
                is_merge_top_left=is_merge_top_left,
                element_type='excel_cell'
            )
            
            # unique_id在__post_init__中自动生成，包含original_index
            
            return element
            
        except Exception as e:
            logger.debug(f"创建精确Excel元素异常: {e}")
            return None

    def _extract_cell_format_precisely(self, cell: Cell) -> Dict[str, Any]:
        """精确提取单元格格式信息"""
        format_info = {}
        try:
            # 字体信息
            if cell.font:
                font = cell.font
                if font.name:
                    format_info['font_name'] = font.name
                if font.size:
                    format_info['font_size'] = font.size
                if font.bold:
                    format_info['bold'] = font.bold
                if font.italic:
                    format_info['italic'] = font.italic
                if font.color and hasattr(font.color, 'rgb') and font.color.rgb:
                    format_info['font_color'] = font.color.rgb
            
            # 填充信息
            if (cell.fill and hasattr(cell.fill, 'start_color') and 
                cell.fill.start_color and hasattr(cell.fill.start_color, 'rgb') and 
                cell.fill.start_color.rgb):
                format_info['fill_color'] = cell.fill.start_color.rgb
            
            # 对齐信息
            if cell.alignment:
                alignment = cell.alignment
                if alignment.horizontal:
                    format_info['horizontal_alignment'] = alignment.horizontal
                if alignment.vertical:
                    format_info['vertical_alignment'] = alignment.vertical
                if alignment.wrap_text:
                    format_info['wrap_text'] = alignment.wrap_text
                if alignment.text_rotation:
                    format_info['text_rotation'] = alignment.text_rotation
            
            # 边框信息
            if cell.border:
                format_info['has_border'] = True
                # 可以进一步提取边框的详细信息
            
            # 数字格式
            if cell.number_format and cell.number_format != 'General':
                format_info['number_format'] = cell.number_format
            
            # 单元格保护
            if cell.protection:
                format_info['protection'] = {
                    'locked': cell.protection.locked,
                    'hidden': cell.protection.hidden
                }
                
        except Exception as e:
            logger.debug(f"提取单元格格式异常: {e}")
        
        return format_info

    def _get_surrounding_context_precisely(self, worksheet: Worksheet, row: int, col: int, radius: int = 1) -> Dict[str, Any]:
        """精确获取周围单元格的上下文信息"""
        context = {}
        try:
            surrounding_values = []
            non_empty_count = 0
            
            for r in range(max(1, row - radius), min(worksheet.max_row + 1, row + radius + 1)):
                for c in range(max(1, col - radius), min(worksheet.max_column + 1, col + radius + 1)):
                    if r != row or c != col:  # 排除自己
                        try:
                            cell_value = worksheet.cell(row=r, column=c).value
                            if cell_value:
                                non_empty_count += 1
                                if len(surrounding_values) < 10:  # 限制数量
                                    surrounding_values.append(str(cell_value)[:50])
                        except:
                            continue
            
            context['surrounding_values'] = surrounding_values
            context['surrounding_count'] = non_empty_count
            context['in_table'] = non_empty_count >= 6  # 如果周围有6个以上非空单元格，认为在表格中
            
            # ✅ 增加位置上下文
            context['is_edge'] = row == 1 or col == 1  # 是否在边缘
            context['relative_position'] = {
                'row': row,
                'col': col,
                'is_top_area': row <= 3,
                'is_left_area': col <= 3
            }
            
        except Exception as e:
            logger.debug(f"获取上下文异常: {e}")
        
        return context

    def _extract_comments_precisely(self, worksheet: Worksheet, sheet_name: str, 
                                  sheet_index: int, start_index: int) -> Tuple[List[SafeExcelElement], int]:
        """精确提取批注"""
        elements = []
        current_index = start_index
        
        try:
            for cell_coordinate, comment in worksheet.comments.items():
                if comment and comment.text:
                    self.stats['comments'] += 1
                    
                    # 解析坐标
                    try:
                        col_letter, row_str = coordinate_from_string(cell_coordinate)
                        row_index = int(row_str)
                        col_index = column_index_from_string(col_letter)
                    except:
                        logger.debug(f"批注坐标解析失败: {cell_coordinate}")
                        continue
                    
                    # ✅ 创建精确映射的批注元素
                    element = SafeExcelElement(
                        full_text=comment.text,
                        cell_value=comment.text,
                        sheet_name=sheet_name,
                        sheet_index=sheet_index,
                        row_index=row_index,
                        col_index=col_index,
                        cell_address=cell_coordinate,
                        original_index=current_index,  # ✅ 精确索引
                        cell_type='comment',
                        comment_info={
                            'author': getattr(comment, 'author', ''),
                            'coordinate': cell_coordinate
                        },
                        element_type='excel_comment'
                    )
                    
                    elements.append(element)
                    current_index += 1
                    self.stats['precise_mappings'] += 1
                    
        except Exception as e:
            logger.debug(f"提取批注异常: {e}")
        
        logger.debug(f"工作表 {sheet_name}: 提取 {len(elements)} 个批注，索引范围 {start_index}-{current_index-1}")
        return elements, current_index

    def _extract_charts_precisely(self, worksheet: Worksheet, sheet_name: str, 
                                sheet_index: int, start_index: int) -> Tuple[List[SafeExcelElement], int]:
        """精确提取图表元素"""
        elements = []
        current_index = start_index
        
        try:
            if hasattr(worksheet, '_charts'):
                for chart_idx, chart in enumerate(worksheet._charts):
                    self.stats['chart_elements'] += 1
                    
                    chart_texts = self._extract_chart_texts_precisely(chart)
                    
                    for text_idx, text in enumerate(chart_texts):
                        # ✅ 创建精确映射的图表元素
                        element = SafeExcelElement(
                            full_text=text,
                            cell_value=text,
                            sheet_name=sheet_name,
                            sheet_index=sheet_index,
                            row_index=1,  # 图表使用虚拟坐标
                            col_index=1,
                            cell_address=f"Chart_{chart_idx}_{text_idx}",
                            original_index=current_index,  # ✅ 精确索引
                            cell_type='chart_element',
                            chart_info={
                                'chart_index': chart_idx,
                                'text_index': text_idx,
                                'chart_type': type(chart).__name__
                            },
                            element_type='excel_chart'
                        )
                        
                        elements.append(element)
                        current_index += 1
                        self.stats['precise_mappings'] += 1
                        
        except Exception as e:
            logger.debug(f"提取图表异常: {e}")
        
        logger.debug(f"工作表 {sheet_name}: 提取 {len(elements)} 个图表元素，索引范围 {start_index}-{current_index-1}")
        return elements, current_index

    def _extract_chart_texts_precisely(self, chart) -> List[str]:
        """精确提取图表文本"""
        texts = []
        try:
            # 尝试提取图表标题
            if hasattr(chart, 'title') and chart.title:
                if hasattr(chart.title, 'text') and chart.title.text:
                    texts.append(chart.title.text)
            
            # 尝试提取轴标签（openpyxl支持有限）
            if hasattr(chart, 'x_axis') and chart.x_axis:
                if hasattr(chart.x_axis, 'title') and chart.x_axis.title:
                    texts.append(chart.x_axis.title)
            
            if hasattr(chart, 'y_axis') and chart.y_axis:
                if hasattr(chart.y_axis, 'title') and chart.y_axis.title:
                    texts.append(chart.y_axis.title)
            
            # 由于openpyxl对图表的支持有限，这里可能需要扩展
        except Exception as e:
            logger.debug(f"提取图表文本异常: {e}")
        
        return texts

    def _get_cache_key(self, text: str, target_lang: str, source_lang: Optional[str], 
                      prompt_config: Optional[Dict[str, Any]] = None) -> str:
        """生成缓存键"""
        try:
            config = prompt_config or self.effective_prompt_config
            
            if self._cached_config_hash is None and config is not None:
                try:
                    mode = config.get('mode', '') if config else ''
                    template = config.get('prompt_template', '') if config else ''
                    custom_system = ''
                    
                    custom_prompt = config.get('custom_prompt')
                    if custom_prompt and isinstance(custom_prompt, dict):
                        custom_system = custom_prompt.get('system', '')[:50]
                    
                    key_config_str = f"{mode}_{template}_{custom_system}"
                    self._cached_config_hash = hashlib.md5(key_config_str.encode('utf-8')).hexdigest()[:8]
                except:
                    self._cached_config_hash = ""
            
            prompt_hash = self._cached_config_hash or ""
            text_hash = hashlib.md5(text.encode('utf-8')).hexdigest()
            return f"{text_hash}_{target_lang}_{source_lang or 'auto'}_{prompt_hash}"
            
        except Exception as e:
            text_hash = hashlib.md5(text.encode('utf-8')).hexdigest()
            return f"{text_hash}_{target_lang}_{source_lang or 'auto'}"

    def _get_enhanced_system_prompt(self, target_lang: str, source_lang: Optional[str]) -> str:
        """获取增强的系统提示"""
        try:
            if not self.effective_prompt_config or self.effective_prompt_config.get('mode') == 'none':
                return self._build_default_excel_system_prompt(target_lang, source_lang)
            
            mode = self.effective_prompt_config.get('mode', 'none')
            
            if mode == 'custom' and self.effective_prompt_config.get('custom_prompt'):
                return self._build_custom_excel_system_prompt(target_lang, source_lang)
            elif mode == 'professional':
                return self._build_professional_excel_system_prompt(target_lang, source_lang)
            elif mode == 'general':
                return self._build_general_excel_system_prompt(target_lang, source_lang)
            else:
                return self._build_simple_excel_system_prompt(target_lang, source_lang)
                
        except Exception as e:
            logger.warning(f"Error generating enhanced system prompt: {e}")
            return self._build_default_excel_system_prompt(target_lang, source_lang)

    def _build_professional_excel_system_prompt(self, target_lang: str, source_lang: Optional[str]) -> str:
        """构建专业Excel模板系统提示"""
        try:
            domain = self.effective_prompt_config.get('prompt_template', 'financial')
            
            professional_prompts = {
                'financial': f"""You are an expert financial spreadsheet translator with expertise in accounting and finance.
Translate from {source_lang or 'auto-detected language'} to {target_lang}.
Maintain financial accuracy, preserve all numerical data, and use standard financial terminology.
Focus on: Financial statement headers, budget labels, financial metrics, accounting terminology, chart of accounts.
Preserve all monetary amounts, percentages, calculations, and account codes exactly as they appear.""",
                
                'data_analysis': f"""You are a data analysis spreadsheet translator specializing in analytical reports.
Translate from {source_lang or 'auto-detected language'} to {target_lang}.
Maintain analytical precision and use appropriate statistical terminology.
Focus on: Data field names, metric definitions, statistical terminology, dashboard labels, KPI descriptions.
Preserve all data values, formulas, calculations, and measurement units exactly as they appear.""",
                
                'inventory': f"""You are an inventory management spreadsheet translator with supply chain expertise.
Translate from {source_lang or 'auto-detected language'} to {target_lang}.
Maintain inventory accuracy and use standard logistics terminology.
Focus on: Product names, inventory status, supply chain terminology, warehouse labels, logistics processes.
Preserve SKU codes, quantities, pricing, barcodes, and product identifiers exactly as they appear.""",
                
                'project': f"""You are a project management spreadsheet translator specializing in project documentation.
Translate from {source_lang or 'auto-detected language'} to {target_lang}.
Maintain project clarity and use standard PM terminology.
Focus on: Task names, milestone descriptions, project phases, resource allocation, deliverable labels.
Preserve dates, durations, progress metrics, and project codes exactly as they appear.""",
                
                'hr': f"""You are an HR spreadsheet translator with expertise in human resources management.
Translate from {source_lang or 'auto-detected language'} to {target_lang}.
Maintain HR confidentiality and use professional HR terminology.
Focus on: Job titles, department names, HR processes, performance metrics, organizational structure.
Preserve employee IDs, dates, salary figures, and confidential codes exactly as they appear.""",
                
                'sales': f"""You are a sales spreadsheet translator with expertise in sales analytics and CRM.
Translate from {source_lang or 'auto-detected language'} to {target_lang}.
Maintain sales accuracy and use standard sales terminology.
Focus on: Customer information, sales stages, revenue metrics, territory management, pipeline data.
Preserve contact information, amounts, dates, and customer codes exactly as they appear."""
            }
            
            system_content = professional_prompts.get(domain, professional_prompts['financial'])
            
            system_content += f"""

PROFESSIONAL EXCEL BATCH PROCESSING RULES:
1. Process each numbered line [1], [2], etc. individually representing different spreadsheet elements
2. Consider the spreadsheet context and maintain professional consistency throughout
3. Keep the exact same number of lines as input - this is critical for precise mapping
4. Output only the translated content, one per line, in the same order
5. Do not include line numbers [1], [2], etc. in your output
6. Preserve all numerical data, formulas, codes, and technical identifiers unchanged
7. Maintain professional Excel terminology and data relationships
8. Do not include any explanatory comments or extra text in your output
9. Ensure translations are contextually appropriate for {target_lang} business environment
10. If a line contains only numbers, codes, or formulas, return it unchanged
11. Handle multi-line content within cells properly by preserving line breaks"""
            
            return self._add_excel_enhancement_rules(system_content)
            
        except Exception as e:
            logger.warning(f"Error building professional Excel system prompt: {e}")
            return self._build_default_excel_system_prompt(target_lang, source_lang)

    def _build_custom_excel_system_prompt(self, target_lang: str, source_lang: Optional[str]) -> str:
        """构建自定义Excel系统提示"""
        try:
            custom_prompt = self.effective_prompt_config.get('custom_prompt', {})
            system_content = custom_prompt.get('system', '')
            
            if not system_content:
                return self._build_default_excel_system_prompt(target_lang, source_lang)
            
            if "numbered line" not in system_content.lower():
                system_content += f"""

EXCEL PRECISION BATCH PROCESSING:
- Each input line is numbered [1], [2], etc. representing different spreadsheet elements
- Translate each numbered line individually while maintaining spreadsheet context
- Keep the exact same number of lines as input for precise element mapping
- Output only the translated content without line numbers
- Preserve all data integrity and Excel functionality
- Ensure translations are appropriate for spreadsheet context in {target_lang}
- Handle multi-line content within cells by preserving internal line breaks"""
            
            return self._add_excel_enhancement_rules(system_content)
            
        except Exception as e:
            return self._build_default_excel_system_prompt(target_lang, source_lang)

    def _build_general_excel_system_prompt(self, target_lang: str, source_lang: Optional[str]) -> str:
        """构建通用Excel增强系统提示"""
        system_content = f"""You are a professional spreadsheet translator with expertise in Excel documents.
Translate from {source_lang or 'auto-detected language'} to {target_lang}.
Provide accurate, natural translations while preserving spreadsheet integrity and data relationships.

GENERAL EXCEL BATCH PROCESSING RULES:
1. Translate each numbered line [1], [2], etc. representing different spreadsheet elements
2. Consider context from surrounding elements for coherent spreadsheet flow
3. Keep the exact same number of lines as input - critical for precise mapping
4. Preserve all formatting, data types, formulas, and structural relationships
5. Output only the translated content, one per line, in exact same order
6. Do not include line numbers [1], [2], etc. in your output
7. Maintain consistency in terminology throughout the spreadsheet
8. Preserve numerical data, dates, codes, and technical identifiers unchanged
9. Do not include any explanatory comments or extra text in your output
10. Ensure natural, professional translations appropriate for {target_lang}
11. Handle multi-line cell content by preserving internal line breaks"""
        
        return self._add_excel_enhancement_rules(system_content)

    def _build_simple_excel_system_prompt(self, target_lang: str, source_lang: Optional[str]) -> str:
        """构建简单Excel系统提示"""
        return f"""Translate spreadsheet content from {source_lang or 'auto-detected language'} to {target_lang}.
Process each numbered line [1], [2], etc. and return the same number of translated lines.
Preserve data integrity and professional terminology. Do not include line numbers in output.
Keep the exact same number of lines for precise mapping. Handle multi-line content properly."""

    def _build_default_excel_system_prompt(self, target_lang: str, source_lang: Optional[str]) -> str:
        """构建默认Excel系统提示"""
        return f"""You are a professional spreadsheet translator. Translate from {source_lang or 'auto-detected language'} to {target_lang}.

EXCEL TRANSLATION RULES:
1. Translate each numbered line individually, considering full spreadsheet context
2. Keep the exact same number of lines as the original for precise element mapping
3. Preserve all formatting, data types, numerical values, and formulas unchanged
4. Output only the translated lines, one per line, in the same order
5. Do not include the original line numbers [1], [2], etc. in your output
6. Do not translate numerical data, formulas, dates, or technical codes
7. Maintain professional Excel terminology and data relationships
8. Return only the translated content without explanatory comments
9. Handle multi-line cell content by preserving internal line breaks"""

    def _add_excel_enhancement_rules(self, system_content: str) -> str:
        """添加Excel特有的增强规则"""
        try:
            if not self.effective_prompt_config or self.effective_prompt_config.get('mode') == 'none':
                return system_content
            
            enhancements = []
            
            preserve_terms = self.effective_prompt_config.get('preserve_terms')
            if preserve_terms and isinstance(preserve_terms, list) and preserve_terms:
                terms = ', '.join(str(term) for term in preserve_terms if term)
                if terms:
                    enhancements.append(f"PRESERVE THESE TERMS EXACTLY: {terms}")
            
            glossary = self.effective_prompt_config.get('glossary')
            if glossary and isinstance(glossary, dict) and glossary:
                try:
                    glossary_text = '; '.join([f"{k}: {v}" for k, v in glossary.items() if k and v])
                    if glossary_text:
                        enhancements.append(f"USE THIS GLOSSARY: {glossary_text}")
                except:
                    pass
            
            additional_context = self.effective_prompt_config.get('additional_context')
            if additional_context and str(additional_context).strip():
                enhancements.append(f"SPREADSHEET CONTEXT: {str(additional_context).strip()}")
            
            if enhancements:
                enhancement_text = "\n\nADDITIONAL EXCEL REQUIREMENTS:\n" + "\n".join(f"• {rule}" for rule in enhancements)
                system_content += enhancement_text
            
            return system_content
            
        except Exception as e:
            return system_content

    def _create_precise_batches(self, elements: List[SafeExcelElement]) -> List[Tuple[List[SafeExcelElement], List[int]]]:
        """创建精确映射批次 - 保持原始顺序，不排序"""
        batches = []
        current_batch = []
        current_original_indices = []  # ✅ 保存原始索引
        current_chars = 0
        
        logger.info(f"创建精确映射批次，总元素数: {len(elements)}")
        
        # ✅ 不排序，保持收集时的原始顺序
        for element in elements:
            text_len = len(element.full_text)
            
            # 检查是否需要开始新批次
            if (len(current_batch) >= self.batch_size or 
                current_chars + text_len > self.max_chars) and current_batch:
                
                batches.append((current_batch, current_original_indices))
                logger.debug(f"批次 {len(batches)}: {len(current_batch)} 个元素, "
                           f"索引范围: {min(current_original_indices)}-{max(current_original_indices)}")
                
                current_batch = [element]
                current_original_indices = [element.original_index]  # ✅ 使用元素的原始索引
                current_chars = text_len
            else:
                current_batch.append(element)
                current_original_indices.append(element.original_index)  # ✅ 使用元素的原始索引
                current_chars += text_len
        
        if current_batch:
            batches.append((current_batch, current_original_indices))
            logger.debug(f"批次 {len(batches)}: {len(current_batch)} 个元素, "
                       f"索引范围: {min(current_original_indices)}-{max(current_original_indices)}")
        
        logger.info(f"创建完成：{len(batches)} 个精确映射批次")
        return batches

    def _translate_batch_with_precise_mapping(self, batch: List[SafeExcelElement], batch_original_indices: List[int],
                                            target_lang: str, source_lang: Optional[str], 
                                            batch_num: int) -> Tuple[bool, List[Tuple[SafeExcelElement, str]], List[bool]]:
        """精确映射的批次翻译 - 确保元素和结果一一对应"""
        try:
            with ThreadPoolExecutor(max_workers=1) as executor:
                future = executor.submit(self._translate_batch_precisely, batch, target_lang, source_lang, batch_num)
                
                try:
                    element_result_pairs, from_cache_flags = future.result(timeout=self.translation_timeout)
                    return True, element_result_pairs, from_cache_flags
                    
                except TimeoutError:
                    logger.warning(f"批次 {batch_num} 翻译超时 ({self.translation_timeout}秒)")
                    # ✅ 返回元素-结果对，确保一一对应
                    timeout_pairs = [(element, f"翻译超时 ({self.translation_timeout}秒)") for element in batch]
                    cache_flags = [False] * len(batch)
                    return False, timeout_pairs, cache_flags
                    
        except Exception as e:
            logger.error(f"批次 {batch_num} 异常: {e}")
            error_pairs = [(element, f"翻译异常: {str(e)}") for element in batch]
            cache_flags = [False] * len(batch)
            return False, error_pairs, cache_flags

    # ✅ 修复后的核心翻译方法
    def _translate_batch_precisely(self, batch: List[SafeExcelElement], target_lang: str, 
                                 source_lang: Optional[str], batch_num: int) -> Tuple[List[Tuple[SafeExcelElement, str]], List[bool]]:
        """精确的批次翻译 - 修复多行文本处理"""
        
        # ✅ 第一阶段：创建元素到文本的精确映射
        element_mapping = {}  # unique_id -> mapping_info
        texts_to_translate = []
        
        local_prompt_config = self._get_config_safe()
        
        logger.debug(f"批次 {batch_num}: 开始精确映射翻译，元素数: {len(batch)}")
        
        # ✅ 建立精确映射，检查缓存
        for batch_index, element in enumerate(batch):
            element_id = element.unique_id
            original_text = element.full_text
            
            mapping_info = {
                'element': element,
                'text': original_text,
                'escaped_text': _escape_excel_text_for_translation(original_text),  # ✅ 新增：转义文本
                'batch_index': batch_index,
                'original_index': element.original_index,
                'translated': None,
                'from_cache': False,
                'cache_key': None
            }
            
            element_mapping[element_id] = mapping_info
            
            # 检查缓存 - 使用原始文本计算缓存键
            cache_key = self._get_cache_key(original_text, target_lang, source_lang, local_prompt_config)
            mapping_info['cache_key'] = cache_key
            cached = self.cache.get(cache_key)
            
            if cached:
                mapping_info['translated'] = cached
                mapping_info['from_cache'] = True
                self.stats['cache_savings'] += 1
                logger.debug(f"缓存命中: {element.unique_id} -> {cached[:30]}...")
            else:
                texts_to_translate.append({
                    'element_id': element_id,
                    'text': original_text,
                    'escaped_text': mapping_info['escaped_text'],  # ✅ 使用转义后的文本
                    'batch_index': batch_index,
                    'original_index': element.original_index
                })
        
        # ✅ 第二阶段：翻译未缓存的文本
        if texts_to_translate:
            logger.debug(f"批次 {batch_num}: 需要翻译 {len(texts_to_translate)} 个未缓存文本")
            
            try:
                system_prompt = self._get_enhanced_system_prompt(target_lang, source_lang)
                
                # ✅ 构建编号文本，使用转义后的文本防止换行符破坏格式
                numbered_texts = [f"[{item['batch_index']+1}] {item['escaped_text']}" for item in texts_to_translate]
                user_message = "\n".join(numbered_texts)
                
                logger.debug(f"批次 {batch_num}: 发送翻译请求，字符数: {len(user_message)}")
                logger.debug(f"批次 {batch_num}: 包含转义标记数量: 换行{user_message.count(NEWLINE_PLACEHOLDER)}, 制表符{user_message.count(TAB_PLACEHOLDER)}")
                
                # 调用翻译API
                translated_result = self._call_translator_safely(user_message, target_lang, source_lang, local_prompt_config, system_prompt)
                self.stats['api_calls'] += 1
                
                # ✅ 解析翻译结果 - 改进的解析逻辑
                translated_parts = self._extract_numbered_translations_enhanced(translated_result, len(texts_to_translate))
                
                logger.debug(f"批次 {batch_num}: 收到翻译结果 {len(translated_parts)} 行")
                
                # ✅ 第三阶段：将翻译结果精确映射回元素
                for i, item in enumerate(texts_to_translate):
                    element_id = item['element_id']
                    mapping_info = element_mapping[element_id]
                    
                    if i < len(translated_parts) and translated_parts[i]:
                        translation = translated_parts[i].strip()
                        # ✅ 还原转义字符
                        translation = _unescape_excel_text_from_translation(translation)
                        mapping_info['translated'] = translation
                        
                        # 更新缓存 - 使用还原后的翻译结果
                        self.cache.put(mapping_info['cache_key'], translation)
                        logger.debug(f"翻译成功: {element_id} -> {translation[:30]}...")
                    else:
                        # 翻译失败，使用原文
                        mapping_info['translated'] = item['text']
                        logger.debug(f"翻译失败，使用原文: {element_id}")
                        
            except Exception as e:
                logger.error(f"批次 {batch_num} 翻译失败: {e}")
                # 翻译失败，所有未缓存的使用原文
                for item in texts_to_translate:
                    element_id = item['element_id']
                    element_mapping[element_id]['translated'] = item['text']
        
        # ✅ 第四阶段：按批次原始顺序构建结果
        element_result_pairs = []
        from_cache_flags = []
        
        for element in batch:  # ✅ 按批次中的原始顺序
            element_id = element.unique_id
            mapping_info = element_mapping[element_id]
            
            translated_text = mapping_info['translated'] or element.full_text
            element_result_pairs.append((element, translated_text))
            from_cache_flags.append(mapping_info['from_cache'])
        
        logger.debug(f"批次 {batch_num}: 精确映射完成，返回 {len(element_result_pairs)} 个结果")
        return element_result_pairs, from_cache_flags

    # ✅ 修复后的解析方法
    def _extract_numbered_translations_enhanced(self, response: str, expected_count: int) -> List[str]:
        """提取编号翻译结果 - 增强版本，处理多行文本"""
        translations = [""] * expected_count
        
        try:
            # ✅ 预处理：识别可能的多行内容
            # 先按编号标记分割，而不是简单按行分割
            response_clean = response.strip()
            
            # 使用正则表达式找到所有编号位置
            number_positions = []
            for match in re.finditer(r'\[(\d+)\]', response_clean):
                line_num = int(match.group(1)) - 1
                start_pos = match.start()
                if 0 <= line_num < expected_count:
                    number_positions.append((line_num, start_pos, match.end()))
            
            # 按位置排序
            number_positions.sort(key=lambda x: x[1])
            
            logger.debug(f"找到 {len(number_positions)} 个编号标记，期望 {expected_count} 个")
            
            # ✅ 提取每个编号对应的内容
            for i, (line_num, start_pos, end_pos) in enumerate(number_positions):
                # 确定内容的结束位置
                if i + 1 < len(number_positions):
                    content_end = number_positions[i + 1][1]  # 下一个编号的开始位置
                else:
                    content_end = len(response_clean)  # 最后一个到字符串末尾
                
                # 提取内容并清理
                content = response_clean[end_pos:content_end].strip()
                
                # 移除可能的前导空白和换行
                content = re.sub(r'^\s+', '', content)
                content = re.sub(r'\s+$', '', content)
                
                if content and 0 <= line_num < expected_count:
                    translations[line_num] = content
                    logger.debug(f"✅ 解析编号行 [{line_num+1}]: {content[:50]}...")
            
            # ✅ 后备处理：如果上述方法没有完全成功，尝试传统行分割
            empty_slots = [i for i, t in enumerate(translations) if not t]
            if empty_slots:
                logger.debug(f"尝试后备解析，还有 {len(empty_slots)} 个空槽")
                
                lines = response_clean.split('\n')
                backup_translations = []
                
                for line in lines:
                    line = line.strip()
                    if not line:
                        continue
                    
                    # 检查是否有编号
                    number_match = NUMBERED_LINE_PATTERN.match(line)
                    if number_match:
                        content = number_match.group(2).strip()
                        if content:
                            backup_translations.append(content)
                    elif not any(line.startswith(f'[{i+1}]') for i in range(expected_count)):
                        # 没有编号的行，可能是多行内容的一部分
                        if backup_translations:
                            # 追加到最后一个翻译
                            backup_translations[-1] += "\n" + line
                
                # 填充空槽
                for i, slot_index in enumerate(empty_slots):
                    if i < len(backup_translations):
                        translations[slot_index] = backup_translations[i]
                        logger.debug(f"✅ 后备填充 [{slot_index+1}]: {backup_translations[i][:50]}...")
            
            filled_count = sum(1 for t in translations if t)
            logger.debug(f"翻译解析完成: {filled_count}/{expected_count} 行")
            
            # ✅ 验证结果完整性
            if filled_count < expected_count:
                logger.warning(f"解析结果不完整: {filled_count}/{expected_count}")
                # 对于空的位置，使用占位符
                for i, translation in enumerate(translations):
                    if not translation:
                        translations[i] = f"[翻译解析失败_{i+1}]"
                        logger.warning(f"位置 {i+1} 解析失败，使用占位符")
            
        except Exception as e:
            logger.error(f"Error extracting numbered translations: {e}")
            # 发生异常时，返回占位符数组
            translations = [f"[解析异常_{i+1}]" for i in range(expected_count)]
        
        return translations

    def _call_translator_safely(self, user_message: str, target_lang: str, source_lang: Optional[str], 
                              local_prompt_config: Dict, system_prompt: str) -> str:
        """安全调用翻译器"""
        try:
            return self.translator.translate(
                text=user_message,
                target_lang=target_lang,
                source_lang=source_lang,
                prompt_config=local_prompt_config,
                config_merge_mode='merge'
            )
        except (TypeError, AttributeError):
            try:
                messages = [
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_message}
                ]
                return self.translator.translate(
                    messages=messages,
                    target_lang=target_lang,
                    source_lang=source_lang
                )
            except:
                full_prompt = f"{system_prompt}\n\nSpreadsheet content to translate:\n{user_message}"
                return self.translator.translate(full_prompt, target_lang, source_lang)

    def _apply_translations_precisely(self, element_result_pairs: List[Tuple[SafeExcelElement, str]], 
                                    from_cache_flags: List[bool]):
        """精确应用翻译结果 - 多重验证"""
        success_count = 0
        validation_failures = 0
        
        logger.debug(f"开始精确应用 {len(element_result_pairs)} 个翻译结果")
        
        for i, (element, translated_text) in enumerate(element_result_pairs):
            try:
                from_cache = from_cache_flags[i] if i < len(from_cache_flags) else False
                
                # ✅ 验证翻译质量
                is_good, reason, score = TranslationValidator.validate_translation_quality(
                    element.full_text, translated_text
                )
                
                if not is_good:
                    validation_failures += 1
                    self.stats['validation_failures'] += 1
                    logger.debug(f"翻译质量问题 {element.unique_id}: {reason} (分数: {score:.2f})")
                
                # ✅ 检查是否为严重失败
                is_serious, failure_reason = TranslationValidator.is_serious_failure(
                    element.full_text, translated_text, self.large_text_threshold, from_cache
                )
                
                if is_serious:
                    self._add_failed_task_precise(element, translated_text, from_cache)
                
                # ✅ 应用翻译到精确位置
                if self._apply_translation_to_precise_location(element, translated_text):
                    success_count += 1
                    
            except Exception as e:
                logger.error(f"应用翻译失败 {element.unique_id}: {e}")
                self._add_failed_task_precise(element, f"应用失败: {str(e)}", False)
        
        logger.info(f"翻译应用完成: 成功 {success_count}/{len(element_result_pairs)}, 质量问题 {validation_failures}")

    def _apply_translation_to_precise_location(self, element: SafeExcelElement, translated_text: str) -> bool:
        """应用翻译到精确位置 - 四重验证系统"""
        if not translated_text or translated_text == element.full_text:
            return True
        
        try:
            # ✅ 第一重验证：工作表存在性
            if element.sheet_name not in self.workbook.sheetnames:
                logger.error(f"工作表不存在: {element.sheet_name}")
                self.stats['coordinate_errors'] += 1
                return False
            
            worksheet = self.workbook[element.sheet_name]
            
            # ✅ 第二重验证：坐标范围合理性
            if (element.row_index < 1 or element.col_index < 1 or
                element.row_index > 1048576 or element.col_index > 16384):  # Excel限制
                logger.error(f"坐标超出Excel范围: {element.unique_id}")
                self.stats['coordinate_errors'] += 1
                return False
            
            # ✅ 第三重验证：获取单元格并验证坐标一致性
            cell = worksheet.cell(row=element.row_index, column=element.col_index)
            if cell.coordinate != element.cell_address:
                logger.error(f"坐标不匹配: 期望{element.cell_address}, 实际{cell.coordinate}")
                self.stats['coordinate_errors'] += 1
                return False
            
            # ✅ 第四重验证：合并单元格检查
            if element.is_merged and not element.is_merge_top_left:
                logger.debug(f"跳过合并单元格非左上角: {element.cell_address}")
                return True
            
            # ✅ 应用翻译根据元素类型
            if element.element_type == 'excel_cell':
                # 保存原始值类型
                original_type = type(element.cell_value)
                
                # 应用翻译
                cell.value = translated_text
                
                # 恢复格式
                self._apply_cell_format_precisely(cell, element.format_info)
                
            elif element.element_type == 'excel_comment':
                # 处理批注
                if element.cell_address in worksheet.comments:
                    worksheet.comments[element.cell_address].text = translated_text
                else:
                    logger.warning(f"批注不存在: {element.cell_address}")
                    return False
            
            elif element.element_type == 'excel_chart':
                # 图表元素处理（有限支持）
                logger.debug(f"图表元素翻译: {element.unique_id}")
            
            self.stats['translated_cells'] += 1
            logger.debug(f"✅ 精确应用成功 {element.unique_id}: {translated_text[:30]}...")
            return True
            
        except Exception as e:
            logger.error(f"❌ 精确应用失败 {element.unique_id}: {e}")
            self.stats['coordinate_errors'] += 1
            return False

    def _apply_cell_format_precisely(self, cell: Cell, format_info: Dict[str, Any]):
        """精确应用单元格格式"""
        try:
            # 应用字体格式
            if any(key in format_info for key in ['font_name', 'font_size', 'bold', 'italic', 'font_color']):
                font_kwargs = {}
                if 'font_name' in format_info:
                    font_kwargs['name'] = format_info['font_name']
                if 'font_size' in format_info:
                    font_kwargs['size'] = format_info['font_size']
                if 'bold' in format_info:
                    font_kwargs['bold'] = format_info['bold']
                if 'italic' in format_info:
                    font_kwargs['italic'] = format_info['italic']
                if 'font_color' in format_info:
                    font_kwargs['color'] = format_info['font_color']
                
                if font_kwargs:
                    cell.font = Font(**font_kwargs)
            
            # 应用填充格式
            if 'fill_color' in format_info:
                cell.fill = PatternFill(start_color=format_info['fill_color'], 
                                      end_color=format_info['fill_color'], 
                                      fill_type='solid')
            
            # 应用对齐格式
            if any(key in format_info for key in ['horizontal_alignment', 'vertical_alignment', 'wrap_text', 'text_rotation']):
                alignment_kwargs = {}
                if 'horizontal_alignment' in format_info:
                    alignment_kwargs['horizontal'] = format_info['horizontal_alignment']
                if 'vertical_alignment' in format_info:
                    alignment_kwargs['vertical'] = format_info['vertical_alignment']
                if 'wrap_text' in format_info:
                    alignment_kwargs['wrap_text'] = format_info['wrap_text']
                if 'text_rotation' in format_info:
                    alignment_kwargs['text_rotation'] = format_info['text_rotation']
                
                if alignment_kwargs:
                    cell.alignment = Alignment(**alignment_kwargs)
            
            # 应用数字格式
            if 'number_format' in format_info:
                cell.number_format = format_info['number_format']
            
            # 应用保护设置
            if 'protection' in format_info:
                from openpyxl.styles import Protection
                protection_info = format_info['protection']
                cell.protection = Protection(
                    locked=protection_info.get('locked', True),
                    hidden=protection_info.get('hidden', False)
                )
                
        except Exception as e:
            logger.debug(f"应用格式失败: {e}")

    def _translate_all_batches_precisely(self, batches: List[Tuple[List[SafeExcelElement], List[int]]], 
                                       target_lang: str, source_lang: Optional[str]):
        """精确翻译所有批次"""
        
        total_elements = sum(len(batch[0]) for batch in batches)
        
        # 根据批次数量决定是否并发
        if len(batches) > 8 and self.max_workers > 1:
            logger.info(f"启用并发翻译: {len(batches)} 个批次, {self.max_workers} 个worker")
            self._translate_concurrent_precisely(batches, target_lang, source_lang)
        else:
            logger.info(f"使用串行翻译: {len(batches)} 个批次")
            self._translate_sequential_precisely(batches, target_lang, source_lang)

    def _translate_sequential_precisely(self, batches: List[Tuple[List[SafeExcelElement], List[int]]], 
                                      target_lang: str, source_lang: Optional[str]):
        """串行精确翻译"""
        with tqdm(total=sum(len(batch[0]) for batch in batches), 
                 desc="Excel精确翻译", unit="单元格") as pbar:
            
            for batch_idx, (batch_elements, batch_original_indices) in enumerate(batches):
                try:
                    # ✅ 精确映射翻译
                    success, element_result_pairs, from_cache_flags = self._translate_batch_with_precise_mapping(
                        batch_elements, batch_original_indices, target_lang, source_lang, batch_idx + 1
                    )
                    
                    if success:
                        # ✅ 精确应用翻译结果
                        self._apply_translations_precisely(element_result_pairs, from_cache_flags)
                    else:
                        # 批次失败，记录所有元素为失败
                        for element, error_msg in element_result_pairs:
                            self._add_failed_task_precise(element, error_msg, False)
                    
                    pbar.update(len(batch_elements))
                    
                except Exception as e:
                    logger.error(f"批次 {batch_idx + 1} 处理异常: {e}")
                    for element in batch_elements:
                        self._add_failed_task_precise(element, f"批次异常: {str(e)}", False)
                    pbar.update(len(batch_elements))

    def _translate_concurrent_precisely(self, batches: List[Tuple[List[SafeExcelElement], List[int]]], 
                                      target_lang: str, source_lang: Optional[str]):
        """并发精确翻译"""
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            with tqdm(total=sum(len(batch[0]) for batch in batches), 
                     desc="Excel精确翻译", unit="单元格") as pbar:
                
                future_to_batch = {}
                for batch_idx, (batch_elements, batch_original_indices) in enumerate(batches):
                    future = executor.submit(
                        self._translate_batch_with_precise_mapping, 
                        batch_elements, batch_original_indices,
                        target_lang, source_lang, batch_idx + 1
                    )
                    future_to_batch[future] = (batch_elements, batch_original_indices, batch_idx + 1)
                
                for future in as_completed(future_to_batch):
                    batch_elements, batch_original_indices, batch_num = future_to_batch[future]
                    try:
                        success, element_result_pairs, from_cache_flags = future.result()
                        
                        if success:
                            self._apply_translations_precisely(element_result_pairs, from_cache_flags)
                        else:
                            for element, error_msg in element_result_pairs:
                                self._add_failed_task_precise(element, error_msg, False)
                        
                        pbar.update(len(batch_elements))
                        
                    except Exception as e:
                        logger.error(f"并发批次 {batch_num} 处理异常: {e}")
                        for element in batch_elements:
                            self._add_failed_task_precise(element, f"并发异常: {str(e)}", False)
                        pbar.update(len(batch_elements))

    def _add_failed_task_precise(self, element: SafeExcelElement, reason: str, from_cache: bool):
        """精确添加失败任务"""
        is_serious, detailed_reason = TranslationValidator.is_serious_failure(
            element.full_text, reason, self.large_text_threshold, from_cache
        )
        
        with self.failed_tasks_lock:
            failed_task = FailedTask(
                original_text=element.full_text,
                original_index=element.original_index,  # ✅ 使用精确的原始索引
                element=element,
                failure_reason=FailureReason.COORDINATE_ERROR if "坐标" in detailed_reason else FailureReason.API_ERROR,
                error_message=detailed_reason,
                is_serious=is_serious
            )
            self.failed_tasks.append(failed_task)
            
            if is_serious:
                self.stats['serious_failures'] += 1
            else:
                self.stats['minor_issues'] += 1

    def translate_excel_precisely(self, input_filepath: str, output_filepath: str, 
                                target_lang: str, source_lang: Optional[str] = None) -> str:
        """主翻译方法 - 精确映射版本"""
        try:
            start_time = time.time()
            self.source_lang = source_lang
            
            # 验证输入文件
            is_valid, validation_msg = _validate_excel_file(input_filepath)
            if not is_valid:
                raise ValueError(f"Excel文件验证失败: {validation_msg}")
            
            with self._translator_config_context():
                logger.info(f"开始Excel精确映射翻译: {os.path.basename(input_filepath)}")
                logger.info(f"目标语言: {target_lang}")
                logger.info(f"配置: batch_size={self.batch_size}, workers={self.max_workers}, min_alpha_ratio={self.min_alpha_ratio}")
                logger.info(validation_msg)
                
                # 复制并加载文件
                shutil.copy2(input_filepath, output_filepath)
                self.workbook = load_workbook(output_filepath)
                
                try:
                    # ✅ 精确收集元素
                    elements = self._collect_excel_elements_precisely()
                    
                    if not elements:
                        logger.info("没有需要翻译的内容")
                        print("翻译完成！（无需要翻译的内容）")
                        return output_filepath
                    
                    self.stats['total_chars'] = sum(len(e.full_text) for e in elements)
                    
                    # ✅ 创建精确映射批次
                    batches = self._create_precise_batches(elements)
                    self.stats['total_batches'] = len(batches)
                    
                    logger.info(f"需要翻译 {len(elements)} 个元素，分为 {len(batches)} 个精确映射批次")
                    
                    # ✅ 清空失败任务
                    with self.failed_tasks_lock:
                        self.failed_tasks.clear()
                    
                    # ✅ 执行精确映射翻译
                    self._translate_all_batches_precisely(batches, target_lang, source_lang)
                    
                    # 保存文件
                    self.workbook.save(output_filepath)
                    
                    self.stats['processing_time'] = time.time() - start_time
                    self._print_stats_precisely()
                    
                    return output_filepath
                    
                finally:
                    if self.workbook:
                        self.workbook.close()
                        self.workbook = None
                        
        except Exception as e:
            logger.error(f"Excel精确映射翻译失败: {e}")
            raise

    def _print_stats_precisely(self):
        """精确统计信息输出"""
        try:
            cache_stats = self.cache.stats
            
            print(f"\n🎯 Excel精确映射翻译完成！")
            print(f"⏱️  处理时间: {self.stats['processing_time']:.1f} 秒")
            print(f"📊 工作表: {self.stats['processed_sheets']}/{self.stats['total_sheets']} 张")
            print(f"📄 单元格: 翻译 {self.stats['translated_cells']} / 跳过 {self.stats['skipped_cells']} / 总计 {self.stats['total_cells']}")
            print(f"📝 处理: {self.stats['total_chars']:,} 字符，{self.stats['total_batches']} 个批次")
            
            # Excel特有统计
            excel_elements = []
            if self.stats['headers'] > 0:
                excel_elements.append(f"表头: {self.stats['headers']}")
            if self.stats['data_cells'] > 0:
                excel_elements.append(f"数据: {self.stats['data_cells']}")
            if self.stats['comments'] > 0:
                excel_elements.append(f"批注: {self.stats['comments']}")
            if self.stats['merge_cell_handled'] > 0:
                excel_elements.append(f"合并单元格: {self.stats['merge_cell_handled']}")
            
            if excel_elements:
                print(f"📋 Excel元素: {', '.join(excel_elements)}")
            
            # 精确映射统计
            print(f"🎯 精确映射: {self.stats['precise_mappings']} 个元素成功映射")
            
            # 智能检测统计
            if self.smart_detection:
                print(f"🔍 智能检测: 通过 {self.stats['smart_detections']} / 过滤 {self.stats['content_filters']}")
            
            # API和缓存
            print(f"🔗 API调用: {self.stats['api_calls']} 次")
            print(f"💾 缓存: 大小 {cache_stats['size']}/{cache_stats['max_size']}, 命中率 {cache_stats['hit_rate']:.1%}, 利用率 {cache_stats['utilization']:.1%}")
            print(f"💾 缓存节省: {self.stats['cache_savings']} 次调用")
            
            # 问题统计
            total_issues = self.stats['serious_failures'] + self.stats['minor_issues'] + self.stats['coordinate_errors'] + self.stats['validation_failures']
            if total_issues > 0:
                print(f"\n⚠️  问题分析:")
                if self.stats['serious_failures'] > 0:
                    print(f"   严重失败: {self.stats['serious_failures']} 个")
                if self.stats['minor_issues'] > 0:
                    print(f"   轻微问题: {self.stats['minor_issues']} 个")
                if self.stats['coordinate_errors'] > 0:
                    print(f"   坐标错误: {self.stats['coordinate_errors']} 个")
                if self.stats['validation_failures'] > 0:
                    print(f"   验证失败: {self.stats['validation_failures']} 个")
            else:
                print("✅ 无翻译问题")
            
            # 配置信息
            print(f"\n⚙️  配置: {self.stats['prompt_mode']} 模式")
            print(f"🔧 参数: batch_size={self.batch_size}, workers={self.max_workers}, min_alpha_ratio={self.min_alpha_ratio}")
            print(f"🔧 阈值: 大段文字={self.large_text_threshold}, 最大长度={self.max_text_length}")
            
            if self.selected_sheets:
                print(f"📋 选择工作表: {', '.join(self.selected_sheets)}")
            
        except Exception as e:
            logger.warning(f"统计信息输出异常: {e}")


def translate_excel_file_formatted(
    input_filepath: str,
    output_dir: str,
    target_lang: str,
    translator,
    source_lang: Optional[str] = None,
    unique_filename_base: Optional[str] = None,
    max_chunk_size: int = 4000,                  # 适应更大批次
    batch_size: int = 50,                        # 用户要求的50
    max_workers: int = 5,                        # 用户要求的5
    retry_max_workers: int = 5,
    reference_doc: Optional[str] = None,
    prompt_config: Optional[Dict[str, Any]] = None,
    translation_timeout: int = 35,               # 适当增加
    max_retries: int = 5,
    large_text_threshold: int = 12,              # 用户调整到12
    retry_failure_threshold: float = 0.0,
    # Excel特有参数
    translate_headers: bool = True,
    translate_comments: bool = True,
    translate_charts: bool = True,
    skip_formulas: bool = True,
    skip_numbers: bool = True,
    skip_dates: bool = True,
    max_text_length: int = 200,
    header_detection_rows: int = 3,              # 减少到3
    min_alpha_ratio: float = 0.2,                # 用户要求的0.2
    selected_sheets: Optional[List[str]] = None,
    preserve_formulas: bool = True,
    preserve_data_validation: bool = True,
    smart_detection: bool = True,
    **kwargs
) -> str:
    """
    高级格式化Excel翻译函数 - 精确映射版本
    
    特点:
    - 精确的元素-结果映射系统
    - 四重坐标验证机制
    - 增强的质量检测
    - 完整的格式保持
    - 智能内容过滤
    - 多行文本正确处理
    """
    
    if not os.path.exists(input_filepath):
        return f"Error: 输入文件未找到: {input_filepath}"
    
    # 验证Excel文件
    is_valid, validation_msg = _validate_excel_file(input_filepath)
    if not is_valid:
        return f"Error: {validation_msg}"
    
    os.makedirs(output_dir, exist_ok=True)
    
    # 生成输出文件路径
    input_path = Path(input_filepath)
    if unique_filename_base:
        output_filename = f"{unique_filename_base}_translated_{target_lang}.xlsx"
    else:
        output_filename = f"{input_path.stem}_translated_{target_lang}.xlsx"
    
    output_filepath = os.path.join(output_dir, output_filename)
    
    # 确保文件名唯一
    counter = 1
    while os.path.exists(output_filepath):
        if unique_filename_base:
            output_filename = f"{unique_filename_base}_translated_{target_lang}_{counter}.xlsx"
        else:
            output_filename = f"{input_path.stem}_translated_{target_lang}_{counter}.xlsx"
        output_filepath = os.path.join(output_dir, output_filename)
        counter += 1

    try:
        # 处理prompt配置
        effective_prompt_config = _prepare_prompt_config(prompt_config, kwargs)
        
        # 获取批处理设置
        batch_settings = _get_batch_settings_from_config(effective_prompt_config, kwargs)
        
        logger.info("=== 启动Excel翻译（精确映射版本 + 多行文本修复）===")
        logger.info(f"输入: {os.path.basename(input_filepath)}")
        logger.info(f"目标语言: {target_lang}")
        logger.info(f"验证结果: {validation_msg}")
        logger.info(f"批次设置: size={batch_settings['batch_size']}, workers={batch_settings['max_workers']}")
        logger.info(f"过滤设置: min_alpha_ratio={min_alpha_ratio}, 智能检测={'启用' if smart_detection else '禁用'}")
        
        if effective_prompt_config and effective_prompt_config.get('mode') != 'none':
            mode = effective_prompt_config.get('mode', 'none')
            logger.info(f"Prompt配置: mode={mode}")
            if mode == 'professional':
                logger.info(f"专业领域: {effective_prompt_config.get('prompt_template', 'financial')}")
        
        # 使用精确映射翻译器
        excel_translator = AdvancedExcelTranslator(
            translator=translator,
            batch_size=batch_settings['batch_size'],
            max_chars=batch_settings['max_chars'],
            max_workers=batch_settings['max_workers'],
            retry_max_workers=retry_max_workers,
            prompt_config=effective_prompt_config,
            reference_doc=reference_doc,
            translation_timeout=batch_settings['translation_timeout'],
            max_retries=batch_settings['max_retries'],
            large_text_threshold=batch_settings['large_text_threshold'],
            retry_failure_threshold=batch_settings['retry_failure_threshold'],
            non_ascii_threshold=batch_settings['non_ascii_threshold'],
            translate_headers=translate_headers,
            translate_comments=translate_comments,
            translate_charts=translate_charts,
            skip_formulas=skip_formulas,
            skip_numbers=skip_numbers,
            skip_dates=skip_dates,
            max_text_length=max_text_length,
            header_detection_rows=header_detection_rows,
            min_alpha_ratio=min_alpha_ratio,
            selected_sheets=selected_sheets,
            preserve_formulas=preserve_formulas,
            preserve_data_validation=preserve_data_validation,
            smart_detection=smart_detection,
            **kwargs
        )
        
        result_path = excel_translator.translate_excel_precisely(
            input_filepath=input_filepath,
            output_filepath=output_filepath,
            target_lang=target_lang,
            source_lang=source_lang
        )

        if not os.path.exists(result_path):
            return "Error: 输出文件未创建"
        
        file_size = os.path.getsize(result_path)
        if file_size == 0:
            return "Error: 输出文件为空"
        
        logger.info(f"Excel精确映射翻译成功完成！输出: {result_path}")
        return result_path
        
        
    except Exception as e:
        error_msg = f"Excel精确映射翻译失败: {str(e)}"
        logger.error(error_msg, exc_info=True)
        return f"Error: {error_msg}"


# 测试用翻译器
class MockExcelTranslator:
    def __init__(self):
        self.call_count = 0
        self.prompt_config = None
    
    def set_prompt_config(self, prompt_config):
        self.prompt_config = copy.deepcopy(prompt_config) if prompt_config else None
    
    def translate(self, text: str = None, target_lang: str = None, 
                 source_lang: str = None, messages: List[Dict] = None, 
                 prompt_config: Optional[Dict[str, Any]] = None,
                 config_merge_mode: str = 'merge', **kwargs) -> str:
        self.call_count += 1
        
        if messages:
            user_content = messages[-1]["content"]
        elif text:
            user_content = text
        else:
            return "Error: No input provided"
        
        # Excel翻译映射
        translation_map = {
            "产品名称": "Product Name", "销售额": "Sales Amount", "数量": "Quantity",
            "单价": "Unit Price", "总计": "Total", "小计": "Subtotal",
            "客户": "Customer", "供应商": "Supplier", "库存": "Inventory",
            "日期": "Date", "金额": "Amount", "状态": "Status",
            "部门": "Department", "员工": "Employee", "项目": "Project",
            "收入": "Revenue", "支出": "Expense", "利润": "Profit",
            "预算": "Budget", "实际": "Actual", "差异": "Variance"
        }
        
        lines = user_content.split('\n')
        translated_lines = []
        
        for line in lines:
            if '[' in line and ']' in line:
                match = re.search(r'\[(\d+)\]\s*(.*)', line)
                if match:
                    content = match.group(2)
                    for zh, en in translation_map.items():
                        content = re.sub(r'\b' + zh + r'\b', en, content)
                    translated_lines.append(content)
        
        return '\n'.join(translated_lines)


if __name__ == '__main__':
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        force=True
    )
    
    print("=== Excel翻译系统 V3.0（精确映射版本）===")
    print("✅ 精确的元素-结果映射系统，彻底解决定位问题")
    print("✅ 四重坐标验证：工作表、范围、地址、合并单元格")
    print("✅ 批次大小: 50，线程数: 5，min_alpha_ratio: 0.2")
    print("✅ 增强的翻译质量验证和评分系统")
    print("✅ 智能内容检测：URL、版本号、配置文件过滤")
    print("✅ 完整的格式保持：字体、填充、对齐、保护")
    print("✅ 增强的缓存系统：1000条目，详细统计")
    print("✅ 专业模板和自定义prompt支持")
    print("✅ 工作表选择性翻译")
    print("✅ 合并单元格正确处理")
    print("✅ Excel文件完整性验证")
    
    print("\n🎯 精确映射核心特性:")
    print("   - original_index: 全局唯一索引")
    print("   - unique_id: 包含索引的精确标识")
    print("   - 元素-结果配对: 一一对应映射")
    print("   - 无排序收集: 保持原始顺序")
    print("   - 多重验证: 四重坐标检查")
    
    print("\n🚀 系统就绪，支持企业级Excel精确翻译！")
